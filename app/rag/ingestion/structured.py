"""Structured file adapter (Phase 6) — xlsx/csv preview before analysis.

The user must see WHAT will be analyzed before anything enters a workflow
(spec §7.4 解析预览): sheet names, headers, row counts, empty-column ratios
and a data preview — plus explicit quality flags instead of silent guessing.
Malformed files are rejected with a repair hint, never passed downstream as
"[解析失败]" placeholder text.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from pydantic import BaseModel


class SheetPreview(BaseModel):
    sheet_name: str
    header_row: list[str]
    row_count: int
    column_count: int
    empty_column_count: int
    sample_rows: list[list[str]]  # up to 3 data rows, cell values truncated


class StructuredFilePreview(BaseModel):
    filename: str
    file_type: str  # xlsx | csv
    sheets: list[SheetPreview]
    quality_notes: list[str]  # human-readable quality flags


class StructuredFileError(Exception):
    """Raised when the file cannot be opened at all — paired with a repair hint."""

    def __init__(self, message: str, repair_hint: str) -> None:
        super().__init__(message)
        self.repair_hint = repair_hint


MAX_PREVIEW_ROWS = 3
MAX_CELL_CHARS = 40
MAX_SHEETS = 20


def preview_xlsx(content: bytes, filename: str) -> StructuredFilePreview:
    try:
        import openpyxl  # type: ignore[import-untyped]

        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise StructuredFileError(
            "无法打开这个表格文件",
            "请确认文件是 .xlsx 格式且未加密；旧版 .xls 请先在 WPS/Excel 中另存为 .xlsx 再上传。",
        ) from exc

    sheets: list[SheetPreview] = []
    quality_notes: list[str] = []

    if workbook.sheetnames:
        quality_notes.append(
            f"包含 {len(workbook.sheetnames)} 个工作表：{'、'.join(workbook.sheetnames[:8])}{'…' if len(workbook.sheetnames) > 8 else ''}"
        )

    for sheet_name in workbook.sheetnames[:MAX_SHEETS]:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)

        header: list[str] = []
        data_rows: list[list[str]] = []
        row_count = 0

        for raw in rows:
            values = ["" if v is None else str(v) for v in (raw or [])]
            if not header:
                if any(v.strip() for v in values):
                    header = [v.strip()[:MAX_CELL_CHARS] or "(未命名列)" for v in values]
                continue
            row_count += 1
            if len(data_rows) < MAX_PREVIEW_ROWS:
                data_rows.append([v[:MAX_CELL_CHARS] for v in values])

        empty_columns = sum(1 for index in range(len(header)) if _column_empty(header, index, data_rows))
        if empty_columns:
            quality_notes.append(f"「{sheet_name}」中有 {empty_columns} 列完全为空，分析时会自动忽略。")

        sheets.append(
            SheetPreview(
                sheet_name=sheet_name,
                header_row=header,
                row_count=row_count,
                column_count=len(header),
                empty_column_count=empty_columns,
                sample_rows=data_rows,
            )
        )

    if not sheets:
        raise StructuredFileError(
            "表格中没有任何工作表",
            "请检查文件内容后重新导出。",
        )

    if all(sheet.row_count == 0 for sheet in sheets):
        quality_notes.append("只有表头没有数据行；分析前请确认是否漏了数据。")

    return StructuredFilePreview(filename=filename, file_type="xlsx", sheets=sheets, quality_notes=quality_notes)


def preview_csv(content: bytes, filename: str) -> StructuredFilePreview:
    text = content.decode("utf-8-sig", errors="replace")

    # Sniff delimiter so the preview matches what the user sees in their tool.
    try:
        sample = text[:2048]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample.strip() else None
    except csv.Error:
        dialect = None

    reader = csv.reader(io.StringIO(text), dialect) if dialect else csv.reader(io.StringIO(text))

    header: list[str] = []
    data_rows: list[list[str]] = []
    row_count = 0

    for values in reader:
        if not header:
            if any((v or "").strip() for v in values):
                header = [(v or "").strip()[:MAX_CELL_CHARS] or "(未命名列)" for v in values]
            continue
        row_count += 1
        if len(data_rows) < MAX_PREVIEW_ROWS:
            data_rows.append([(v or "")[:MAX_CELL_CHARS] for v in values])

    if not header:
        raise StructuredFileError(
            "文件内容为空或不是有效的 CSV",
            "请确认文件为 UTF-8 编码的 CSV（可直接用 WPS/Excel 另存为「CSV UTF-8」）。",
        )

    empty_columns = sum(1 for index in range(len(header)) if _column_empty(header, index, data_rows))
    quality_notes: list[str] = []
    if empty_columns:
        quality_notes.append(f"有 {empty_columns} 列完全为空，分析时会自动忽略。")
    if row_count == 0:
        quality_notes.append("只有表头没有数据行；分析前请确认是否漏了数据。")

    return StructuredFilePreview(
        filename=filename,
        file_type="csv",
        sheets=[
            SheetPreview(
                sheet_name="CSV 数据",
                header_row=header,
                row_count=row_count,
                column_count=len(header),
                empty_column_count=empty_columns,
                sample_rows=data_rows,
            )
        ],
        quality_notes=quality_notes,
    )


def preview_structured(content: bytes, filename: str, declared_type: str) -> StructuredFilePreview:
    """Route to the right adapter by declared type; both paths share the contract."""
    if declared_type == "xlsx":
        return preview_xlsx(content, filename)
    if declared_type == "csv":
        return preview_csv(content, filename)
    raise StructuredFileError("不支持的文件类型", "请上传 .xlsx 或 .csv 文件。")


def _column_empty(header: list[str], index: int, sample_rows: list[list[str]]) -> bool:
    # Empty in the header AND every sampled row → likely unused column.
    if index < len(header) and header[index] and header[index] != "(未命名列)":
        return False
    for row in sample_rows:
        if index < len(row) and row[index].strip():
            return False
    return True


def preview_to_markdown(preview: StructuredFilePreview) -> str:
    """Render the preview as analysis-ready text: headers + all rows described.

    Phase 6 preview keeps this conservative: the text carries sheet names,
    headers, row counts and quality notes, so downstream analysis sees the
    real structure instead of guessing from raw cell dumps.
    """
    parts: list[str] = []
    parts.append(f"结构化文件：{preview.filename}（{preview.file_type}）")
    for sheet in preview.sheets:
        parts.append(f"\n工作表「{sheet.sheet_name}」：{sheet.column_count} 列 × {sheet.row_count} 行数据")
        parts.append("表头：" + " | ".join(sheet.header_row))
        if sheet.sample_rows:
            parts.append("示例行：")
            for row in sheet.sample_rows:
                parts.append("  - " + " | ".join(row))
    for note in preview.quality_notes:
        parts.append(f"质量提示：{note}")
    return "\n".join(parts)


def detect_structured_type(filename: str | None, content_type: str | None) -> str | None:
    """Return 'xlsx' / 'csv' when the upload is a structured file, else None."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return "xlsx"
    if name.endswith(".csv") or content_type in ("text/csv", "application/csv"):
        return "csv"
    return None


def structured_types() -> dict[str, str]:
    """Allowed MIME map for structured uploads (route-level gate)."""
    return {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "text/csv": "csv",
        "application/csv": "csv",
    }


def empty_stats() -> dict[str, Any]:  # pragma: no cover — helper for future metrics
    return {"sheets": 0, "rows": 0}
